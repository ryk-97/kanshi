import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.options import Options
import logging
import os

# ログ設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Excelファイルの設定
excel_file = 'project_votes.xlsx'

# Seleniumのセットアップ
options = webdriver.ChromeOptions()
options.add_argument("--headless=new")
driver = webdriver.Chrome(options=options)
driver.get('https://tasuketsu.com/result/yRQlxbBfWxdBhlO3FVT6')

# 画像をクリックして、必要な要素が表示されるのを待機
try:
    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//img[@src="/_nuxt/img/list2.8ee1fba.png"]'))
    )
    element.click()
except Exception as e:
    logging.error("画像をクリックできませんでした: %s", e)


# Excelファイルが存在しない場合は新規作成
if not os.path.exists(excel_file):
    wb = Workbook()
    sheet = wb.active
    projects = driver.find_elements(By.CSS_SELECTOR, ".w-100 .flex-column .flex-center")
    header = ['Timestamp'] + [project.find_element(By.XPATH, ".//div[contains(@style, 'font-size: 20px')]").text.strip()[2:] for project in projects]
    sheet.append(header)
    wb.save(excel_file)
else:
    wb = load_workbook(excel_file)
    sheet = wb.active
    projects = driver.find_elements(By.CSS_SELECTOR, ".w-100 .flex-column .flex-center")
    header = ['Timestamp'] + [project.find_element(By.XPATH, ".//div[contains(@style, 'font-size: 20px')]").text.strip()[2:] for project in projects]

    # ヘッダーが既に存在するか確認し、存在しない場合は追加
    if sheet.max_row == 0:
        sheet.append(header)

    wb.save(excel_file)

while True:
    # リロードして、ログイン画面が表示されたらログインする
    driver.refresh()
    try:
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//img[@src="/_nuxt/img/list2.8ee1fba.png"]'))
        )
        element.click()
    except Exception as e:
        logging.error("画像をクリックできませんでした: %s", e)
    
    # データを取得して、Excelファイルに書き込む
    projects = driver.find_elements(By.CSS_SELECTOR, ".w-100 .flex-column .flex-center")
    sheet.cell(row=sheet.max_row+1, column=1, value=time.strftime("%Y-%m-%d %H:%M:%S"))
    for project in projects:
        name = project.find_element(By.XPATH, ".//div[contains(@style, 'font-size: 20px')]").text.strip()[2:]
        votes = project.find_element(By.XPATH, ".//div[contains(@style, 'font-size: 16px')]").text.strip()[:-1]
        col_index = header.index(name) + 1
        sheet.cell(row=sheet.max_row, column=col_index, value=int(votes))
    
    wb.save(excel_file)
    
    time.sleep(30)  # 3秒待機
    logging.info("リストを更新しました！")
