import os
import re
import json
import time
import ctypes
import string
import random
import logging
import requests
import threading
from openpyxl import Workbook, load_workbook

session = requests.Session()
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
excel_file = 'project_votes.xlsx'

output = []

url_base = (
    "https://firestore.googleapis.com/google.firestore.v1.Firestore/Listen/channel?"
    "gsessionid={}&VER=8&database=projects%2Ftasuketsu2%2Fdatabases%2F(default)"
    "&RID=rpc&SID={}&CI=0&AID=0&TYPE=xmlhttp&zx={}&t=1"
)

get_gsessionid_url = (
    "https://firestore.googleapis.com/google.firestore.v1.Firestore/Listen/channel?"
    "VER=8&database=projects%2Ftasuketsu2%2Fdatabases%2F(default)"
    "&RID={}&CVER=22&X-HTTP-Session-Id=gsessionid&zx={}&t=1"
)

payload_data = (
    "headers=X-Goog-Api-Client%3Agl-js%2F%20fire%2F9.15.0%0D%0A"
    "Content-Type%3Atext%2Fplain%0D%0AX-Firebase-GMPID%3A1%3A978100370888%3A"
    "web%3A006bc5c6eb78ae899854fd%0D%0A&count=1&ofs=0&req0___data__=%7B%22database%22"
    "%3A%22projects%2Ftasuketsu2%2Fdatabases%2F(default)%22%2C%22addTarget%22%3A%7B"
    "%22documents%22%3A%7B%22documents%22%3A%5B%22projects%2Ftasuketsu2%2Fdatabases"
    "%2F(default)%2Fdocuments%2Fpublic%2FyRQlxbBfWxdBhlO3FVT6%22%5D%7D%2C%22targetId"
    "%22%3A2%7D%7D"
)

def random_string_lowercase(length):
    return ''.join(random.choice(string.ascii_lowercase + string.digits) for _ in range(length))


def random_string(length):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

class CustomThread(threading.Thread):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._run = self.run
        self.run = self.set_id_and_run

    def set_id_and_run(self):
        self.id = threading.get_native_id()
        self._run()

    def get_id(self):
        return self.id
        
    def raise_exception(self):
        res = ctypes.pythonapi.PyThreadState_SetAsyncExc(
            ctypes.c_long(self.get_id()), 
            ctypes.py_object(SystemExit)
        )
        if res > 1:
            ctypes.pythonapi.PyThreadState_SetAsyncExc(
                ctypes.c_long(self.get_id()), 
                0
            )
            print('Failure in raising exception')

def update_output():
    random_int = str(random.randrange(10**4, 10**5))
    
    
    headers = {
        "Content-Length": str(len(payload_data)),
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    gsessionid_response = session.post(
        get_gsessionid_url.format(random_int, random_string_lowercase(5)),
        data=payload_data, headers=headers
    )
    
    if "X-HTTP-Session-Id" not in gsessionid_response.headers:
        print("X-HTTP-Session-Id not found in headers.")
        print(gsessionid_response.headers)
        exit()
    
    get_gsessionid = gsessionid_response.headers["X-HTTP-Session-Id"]
    lines = gsessionid_response.text.splitlines()
    
    try:
        json_data = json.loads(lines[1])
        get_sid = json_data[0][1][1]
    except (IndexError, json.JSONDecodeError) as e:
        print(f"Error parsing JSON: {e}")
        print(lines)
        exit()
       
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "ja;q=0.9",
        "Priority": "u=0, i",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Sec-Gpc": "1",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0 Config/100.2.9281.82"
    }
    
    create_url = url_base.format(get_gsessionid, get_sid, random_string_lowercase(12))
       
    logging.info("データを取得中...")
    
    def temp_request():
        time.sleep(0.3)
        session.get(
            url=create_url,
            headers=headers,
            timeout=1
        )
    
    temp_thread = CustomThread(target=temp_request)
    temp_thread.start()
    
    response = session.get(
        url=create_url,
        headers=headers, timeout=10
    )
    
    temp_thread.raise_exception()
    
    lines = response.text.splitlines()
    
    pattern = r'("documentChange": {.*?"targetIds": \[\s*\d+\s*\]\s*})'
    
    # マッチする部分を抽出
    match = re.search(pattern, response.text, re.DOTALL)
    
    if match:
        extracted_text = match.group(1)
        abc = json.loads("{"+extracted_text+"}")
    else:
        print("該当するテキストが見つかりませんでした。")
        
    # Zの整数値
    data = abc["documentChange"]
    # z-○○の部分を収集
    z_keys = [
        int(key.split('-')[1].split('_')[0])
        for key in data['document']['fields'].keys()
        if key.startswith('z-')
    ]
    
    # データの収集
    for num in z_keys:
        if str(num) in data['document']['fields']['options']['mapValue']['fields']:
            field = data['document']['fields']['options']['mapValue']['fields'][str(num)]
            if 'mapValue' in field and 'fields' in field['mapValue'] and 'name' in field['mapValue']['fields']:
                name = field['mapValue']['fields']['name']['stringValue']
                integer_value = int(data['document']['fields'][f'z-{num}_1']['integerValue'])
                output.append((name, integer_value))
    
    # integerValueでソート
    output.sort(key=lambda x: x[1], reverse=True)

while True:
    output = [] # 初期化
    update_output()
    
    # Excelファイルが存在しない場合は新規作成
    if not os.path.exists(excel_file):
        wb = Workbook()
        sheet = wb.active
        header = ['Timestamp'] + [name[2:] for name, value in output]
        sheet.append(header)
        wb.save(excel_file)
    else:
        wb = load_workbook(excel_file)
        sheet = wb.active
        header = ['Timestamp'] + [name[2:] for name, value in output]
    
        # ヘッダーが既に存在するか確認し、存在しない場合は追加
        if sheet.max_row == 0:
            sheet.append(header)
    
        wb.save(excel_file)
    
    sheet.cell(row=sheet.max_row+1, column=1, value=time.strftime("%Y-%m-%d %H:%M:%S"))
    for name, value in output:
        col_index = header.index(name[2:]) + 1
        sheet.cell(row=sheet.max_row, column=col_index, value=int(value))
    
    wb.save(excel_file)
    
    logging.info("リストを更新しました")
    time.sleep(30) #30秒ごとに更新